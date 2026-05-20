"""Tests for Excel pivot pack output."""

import pytest
from openpyxl import load_workbook

from src.config import REPORTS_DIR


def test_pivot_pack_exists():
    """pivot_pack.xlsx should exist."""
    path = REPORTS_DIR / "pivot_pack.xlsx"
    assert path.exists(), "pivot_pack.xlsx not found"


def test_pivot_pack_has_6_sheets():
    """Pivot pack should have exactly 6 sheets."""
    wb = load_workbook(REPORTS_DIR / "pivot_pack.xlsx")
    assert len(wb.sheetnames) == 6, f"Expected 6 sheets, got {len(wb.sheetnames)}: {wb.sheetnames}"


def test_pivot_pack_sheet_names():
    """Verify expected sheet names."""
    wb = load_workbook(REPORTS_DIR / "pivot_pack.xlsx")
    expected = {"Cover", "Carrier Scorecard", "Lane Heatmap",
                "Shipper Cost-to-Serve", "Accessorial Drivers", "Daily Trend"}
    actual = set(wb.sheetnames)
    assert expected == actual, f"Sheet mismatch. Expected: {expected}, Got: {actual}"


def test_pivot_pack_cover_has_kpis():
    """Cover sheet should have KPI data."""
    wb = load_workbook(REPORTS_DIR / "pivot_pack.xlsx")
    ws = wb["Cover"]
    # Title should be in A1
    assert ws["A1"].value is not None, "Cover sheet A1 is empty"


def test_pivot_pack_carrier_scorecard_has_data():
    """Carrier Scorecard sheet should have header and data rows."""
    wb = load_workbook(REPORTS_DIR / "pivot_pack.xlsx")
    ws = wb["Carrier Scorecard"]
    # Header in row 1
    assert ws.cell(row=1, column=1).value == "Carrier"
    # At least some data rows
    assert ws.max_row > 1, "Carrier Scorecard has no data rows"
