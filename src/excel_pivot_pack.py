"""
Excel pivot pack generator. Produces a 6-sheet pivot_pack.xlsx
with formatted pivot tables for stakeholder delivery.

Usage:
    python -m src.excel_pivot_pack
"""

import logging
from datetime import date

import duckdb
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.config import EXCEL_FONT_NAME, EXCEL_FONT_SIZE, REPORTS_DIR, WAREHOUSE_PATH

logger = logging.getLogger(__name__)

HEADER_FONT = Font(name=EXCEL_FONT_NAME, size=EXCEL_FONT_SIZE, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
DATA_FONT = Font(name=EXCEL_FONT_NAME, size=EXCEL_FONT_SIZE)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
GREEN_FILL = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
RED_FILL = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")


def _write_header(ws, row, headers):
    """Write formatted header row."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _write_data_row(ws, row, values):
    """Write a data row with formatting."""
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        if isinstance(val, float):
            cell.number_format = "#,##0.00"


def _auto_width(ws):
    """Auto-fit column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 30)


def build_cover(wb, con):
    """Sheet 1: Cover page with KPI snapshot."""
    ws = wb.active
    ws.title = "Cover"

    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "Inbound First-Mile Carrier Performance -- Pivot Pack"
    title_cell.font = Font(name=EXCEL_FONT_NAME, size=18, bold=True, color="2C3E50")

    ws["A3"] = f"Generated: {date.today().isoformat()}"
    ws["A4"] = "Dashboard: See Streamlit app (make app) or Tableau Public link in README"
    ws["A3"].font = DATA_FONT
    ws["A4"].font = DATA_FONT

    kpis = con.execute("""
        SELECT
            COUNT(*) AS shipments,
            ROUND(AVG(CASE WHEN on_time_pickup THEN 1.0 ELSE 0.0 END) * 100, 1) AS otp,
            ROUND(AVG(CASE WHEN defect_flag THEN 1.0 ELSE 0.0 END) * 100, 2) AS defect,
            ROUND(AVG(total_cost_usd), 2) AS avg_cost,
            ROUND(AVG(dwell_minutes) FILTER (WHERE dwell_minutes > 0), 1) AS avg_dwell
        FROM fact_shipment
    """).fetchone()

    ws["A6"] = "KPI Snapshot"
    ws["A6"].font = Font(name=EXCEL_FONT_NAME, size=14, bold=True)

    labels = ["Total Shipments", "OTP %", "Defect Rate %", "Avg Cost/Shipment", "Avg Dwell (min)"]
    values = list(kpis)
    for i, (label, val) in enumerate(zip(labels, values)):
        ws.cell(row=7 + i, column=1, value=label).font = Font(name=EXCEL_FONT_NAME, bold=True)
        ws.cell(row=7 + i, column=2, value=val).font = DATA_FONT

    ws.freeze_panes = "A6"


def build_carrier_scorecard(wb, con):
    """Sheet 2: Carrier x Month pivot with OTP, Defect, CPM."""
    ws = wb.create_sheet("Carrier Scorecard")

    data = con.execute("""
        SELECT
            c.carrier_name, c.carrier_tier, d.month,
            ROUND(AVG(CASE WHEN f.on_time_pickup THEN 1.0 ELSE 0.0 END) * 100, 1) AS otp_pct,
            ROUND(AVG(CASE WHEN f.defect_flag THEN 1.0 ELSE 0.0 END) * 100, 2) AS defect_pct,
            ROUND(AVG(f.total_cost_usd / NULLIF(f.distance_miles, 0)), 2) AS cpm
        FROM fact_shipment f
        JOIN dim_carrier c ON f.carrier_id = c.carrier_id
        JOIN dim_date d ON f.shipment_date_key = d.date_key
        GROUP BY c.carrier_name, c.carrier_tier, d.month
        ORDER BY c.carrier_tier, c.carrier_name, d.month
    """).fetchall()

    headers = ["Carrier", "Tier", "Month", "OTP %", "Defect %", "CPM"]
    _write_header(ws, 1, headers)
    for i, row in enumerate(data, 2):
        _write_data_row(ws, i, list(row))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(data) + 1}"
    _auto_width(ws)


def build_lane_heatmap(wb, con):
    """Sheet 3: Origin x Destination OTP with conditional formatting."""
    ws = wb.create_sheet("Lane Heatmap")

    data = con.execute("""
        SELECT
            l.origin_city, l.dest_fc_id,
            ROUND(AVG(CASE WHEN f.on_time_pickup THEN 1.0 ELSE 0.0 END) * 100, 1) AS otp_pct,
            COUNT(*) AS volume
        FROM fact_shipment f
        JOIN dim_lane l ON f.lane_id = l.lane_id
        GROUP BY l.origin_city, l.dest_fc_id
        ORDER BY otp_pct ASC
    """).fetchall()

    headers = ["Origin", "Destination", "OTP %", "Volume"]
    _write_header(ws, 1, headers)
    for i, row in enumerate(data, 2):
        _write_data_row(ws, i, list(row))
        otp_cell = ws.cell(row=i, column=3)
        if row[2] >= 90:
            otp_cell.fill = GREEN_FILL
            otp_cell.font = Font(name=EXCEL_FONT_NAME, size=EXCEL_FONT_SIZE, color="FFFFFF")
        elif row[2] >= 80:
            otp_cell.fill = YELLOW_FILL
        else:
            otp_cell.fill = RED_FILL
            otp_cell.font = Font(name=EXCEL_FONT_NAME, size=EXCEL_FONT_SIZE, color="FFFFFF")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{len(data) + 1}"
    _auto_width(ws)


def build_shipper_cts(wb, con):
    """Sheet 4: Shipper x Service Type cost-to-serve."""
    ws = wb.create_sheet("Shipper Cost-to-Serve")

    data = con.execute("""
        SELECT
            s.shipper_name, s.ship_volume_tier, st.service_name,
            ROUND(SUM(f.total_cost_usd), 2) AS total_cost,
            COUNT(*) AS shipment_count,
            ROUND(AVG(f.total_cost_usd), 2) AS cps
        FROM fact_shipment f
        JOIN dim_shipper s ON f.shipper_id = s.shipper_id
        JOIN dim_service_type st ON f.service_type_id = st.service_type_id
        WHERE s.ship_volume_tier = 'Top20'
        GROUP BY s.shipper_name, s.ship_volume_tier, st.service_name
        ORDER BY total_cost DESC
    """).fetchall()

    headers = ["Shipper", "Volume Tier", "Service Type", "Total Cost", "Shipments", "Cost/Shipment"]
    _write_header(ws, 1, headers)
    for i, row in enumerate(data, 2):
        _write_data_row(ws, i, list(row))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(data) + 1}"
    _auto_width(ws)


def build_accessorial_drivers(wb, con):
    """Sheet 5: Accessorial by reason x month."""
    ws = wb.create_sheet("Accessorial Drivers")

    data = con.execute("""
        SELECT
            COALESCE(f.defect_reason, 'Standard') AS reason,
            d.month,
            ROUND(SUM(f.accessorial_cost_usd), 2) AS total_accessorial,
            COUNT(*) AS shipment_count
        FROM fact_shipment f
        JOIN dim_date d ON f.shipment_date_key = d.date_key
        GROUP BY reason, d.month
        ORDER BY total_accessorial DESC
    """).fetchall()

    headers = ["Reason", "Month", "Total Accessorial $", "Shipment Count"]
    _write_header(ws, 1, headers)
    for i, row in enumerate(data, 2):
        _write_data_row(ws, i, list(row))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{len(data) + 1}"
    _auto_width(ws)


def build_daily_trend(wb, con):
    """Sheet 6: Daily KPI table."""
    ws = wb.create_sheet("Daily Trend")

    data = con.execute("""
        SELECT
            d.date,
            COUNT(*) AS shipments,
            ROUND(AVG(CASE WHEN f.on_time_pickup THEN 1.0 ELSE 0.0 END) * 100, 1) AS otp_pct,
            ROUND(AVG(CASE WHEN f.defect_flag THEN 1.0 ELSE 0.0 END) * 100, 2) AS defect_pct,
            ROUND(AVG(f.total_cost_usd), 2) AS avg_cost,
            ROUND(AVG(f.dwell_minutes) FILTER (WHERE f.dwell_minutes > 0), 1) AS avg_dwell
        FROM fact_shipment f
        JOIN dim_date d ON f.shipment_date_key = d.date_key
        GROUP BY d.date
        ORDER BY d.date DESC
        LIMIT 90
    """).fetchall()

    headers = ["Date", "Shipments", "OTP %", "Defect %", "Avg Cost", "Avg Dwell"]
    _write_header(ws, 1, headers)
    for i, row in enumerate(data, 2):
        _write_data_row(ws, i, list(row))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(data) + 1}"
    _auto_width(ws)


def main():
    logging.basicConfig(
        format="%(asctime)s [%(name)s] %(levelname)s: %(message)s",
        level=logging.INFO,
    )
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    con = duckdb.connect(str(WAREHOUSE_PATH), read_only=True)
    wb = Workbook()

    build_cover(wb, con)
    build_carrier_scorecard(wb, con)
    build_lane_heatmap(wb, con)
    build_shipper_cts(wb, con)
    build_accessorial_drivers(wb, con)
    build_daily_trend(wb, con)

    output_path = REPORTS_DIR / "pivot_pack.xlsx"
    wb.save(output_path)
    con.close()

    logger.info("Pivot pack saved to %s (%d sheets)", output_path, len(wb.sheetnames))


if __name__ == "__main__":
    main()
